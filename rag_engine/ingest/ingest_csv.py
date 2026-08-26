import csv
import io
import os
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from rag_engine.storage.chroma_memory import get_chroma_client

SUPPORTED_EXTENSIONS = {".csv", ".zip", ".xlsx", ".xls"}
DEFAULT_LOCAL_CAR_SPECS_CSV = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "storage", "data", "car details.csv")

# The source dataset has no body-type column, so "SUV"/"sedan"/etc. queries can't
# match anything in the raw text. This maps (Make, base model word) -> body type
# from real-world knowledge of these specific Indian-market model lines, so the
# enriched document text below can actually mention a body type.
CAR_BODY_TYPES: dict[tuple[str, str], str] = {
    ("Audi", "A3"): "Sedan", ("Audi", "A4"): "Sedan", ("Audi", "A6"): "Sedan",
    ("Audi", "A7"): "Coupe", ("Audi", "A8"): "Sedan", ("Audi", "Q2"): "SUV",
    ("Audi", "Q3"): "SUV", ("Audi", "Q5"): "SUV", ("Audi", "Q7"): "SUV",
    ("Audi", "Q8"): "SUV", ("Audi", "RS5"): "Coupe", ("Audi", "TT"): "Coupe",
    ("BMW", "2"): "Coupe", ("BMW", "3-Series"): "Sedan", ("BMW", "5-Series"): "Sedan",
    ("BMW", "6-Series"): "Coupe", ("BMW", "7-Series"): "Sedan", ("BMW", "X1"): "SUV",
    ("BMW", "X3"): "SUV", ("BMW", "X4"): "SUV", ("BMW", "X5"): "SUV",
    ("BMW", "X6"): "SUV", ("BMW", "X7"): "SUV",
    ("Chevrolet", "Beat"): "Hatchback", ("Chevrolet", "Captiva"): "SUV",
    ("Chevrolet", "Cruze"): "Sedan", ("Chevrolet", "Sail"): "Sedan", ("Chevrolet", "Spark"): "Hatchback",
    ("Datsun", "Go"): "Hatchback", ("Datsun", "Redigo"): "Hatchback",
    ("Ferrari", "488"): "Sports Car",
    ("Fiat", "Linea"): "Sedan", ("Fiat", "Punto"): "Hatchback",
    ("Ford", "Aspire"): "Sedan", ("Ford", "Ecosport"): "SUV", ("Ford", "Endeavour"): "SUV",
    ("Ford", "Fiesta"): "Sedan", ("Ford", "Figo"): "Hatchback", ("Ford", "Ikon"): "Sedan",
    ("Ford", "Mustang"): "Coupe",
    ("Honda", "Accord"): "Sedan", ("Honda", "Amaze"): "Sedan", ("Honda", "BR-V"): "SUV",
    ("Honda", "Brio"): "Hatchback", ("Honda", "CR-V"): "SUV", ("Honda", "City"): "Sedan",
    ("Honda", "Civic"): "Sedan", ("Honda", "Jazz"): "Hatchback", ("Honda", "Mobilio"): "MUV/MPV",
    ("Honda", "WR-V"): "SUV",
    ("Hyundai", "Accent"): "Sedan", ("Hyundai", "Alcazar"): "SUV", ("Hyundai", "Aura"): "Sedan",
    ("Hyundai", "Creta"): "SUV", ("Hyundai", "Elantra"): "Sedan", ("Hyundai", "Elite"): "Hatchback",
    ("Hyundai", "Eon"): "Hatchback", ("Hyundai", "Grand"): "Hatchback", ("Hyundai", "Santa"): "SUV",
    ("Hyundai", "Santro"): "Hatchback", ("Hyundai", "Sonata"): "Sedan", ("Hyundai", "Tucson"): "SUV",
    ("Hyundai", "Venue"): "SUV", ("Hyundai", "Verna"): "Sedan", ("Hyundai", "Xcent"): "Sedan",
    ("Hyundai", "i10"): "Hatchback", ("Hyundai", "i20"): "Hatchback",
    ("Isuzu", "MU-X"): "SUV",
    ("Jaguar", "F-Pace"): "SUV", ("Jaguar", "XE"): "Sedan", ("Jaguar", "XF"): "Sedan", ("Jaguar", "XJ"): "Sedan",
    ("Jeep", "Compass"): "SUV", ("Jeep", "Wrangler"): "SUV",
    ("Kia", "Carnival"): "MUV/MPV", ("Kia", "Seltos"): "SUV", ("Kia", "Sonet"): "SUV",
    ("Lamborghini", "Huracan"): "Sports Car",
    ("Land Rover", "Discovery"): "SUV", ("Land Rover", "Evoque"): "SUV", ("Land Rover", "Range"): "SUV",
    ("Lexus", "ES"): "Sedan", ("Lexus", "LX"): "SUV", ("Lexus", "NX"): "SUV",
    ("MG", "Astor"): "SUV", ("MG", "Gloster"): "SUV", ("MG", "Hector"): "SUV", ("MG", "ZS"): "SUV",
    ("MINI", "Cooper"): "Hatchback", ("MINI", "Countryman"): "SUV",
    ("Mahindra", "Alturas"): "SUV", ("Mahindra", "Bolero"): "SUV", ("Mahindra", "KUV100"): "SUV",
    ("Mahindra", "Marazzo"): "MUV/MPV", ("Mahindra", "Quanto"): "SUV", ("Mahindra", "Scorpio"): "SUV",
    ("Mahindra", "TUV300"): "SUV", ("Mahindra", "Thar"): "SUV", ("Mahindra", "XUV300"): "SUV",
    ("Mahindra", "XUV500"): "SUV", ("Mahindra", "XUV700"): "SUV", ("Mahindra", "Xylo"): "MUV/MPV",
    ("Maruti Suzuki", "Alto"): "Hatchback", ("Maruti Suzuki", "Baleno"): "Hatchback",
    ("Maruti Suzuki", "Celerio"): "Hatchback", ("Maruti Suzuki", "Ciaz"): "Sedan",
    ("Maruti Suzuki", "DZire"): "Sedan", ("Maruti Suzuki", "Eeco"): "MUV/MPV",
    ("Maruti Suzuki", "Ertiga"): "MUV/MPV", ("Maruti Suzuki", "Estilo"): "Hatchback",
    ("Maruti Suzuki", "Grand"): "SUV", ("Maruti Suzuki", "Ignis"): "Hatchback",
    ("Maruti Suzuki", "Ritz"): "Hatchback", ("Maruti Suzuki", "S-Cross"): "SUV",
    ("Maruti Suzuki", "S-Presso"): "Hatchback", ("Maruti Suzuki", "SX4"): "Sedan",
    ("Maruti Suzuki", "Swift"): "Hatchback", ("Maruti Suzuki", "Vitara"): "SUV",
    ("Maruti Suzuki", "Wagon"): "Hatchback", ("Maruti Suzuki", "XL6"): "MUV/MPV",
    ("Maruti Suzuki", "Zen"): "Hatchback",
    ("Maserati", "Levante"): "SUV",
    ("Mercedes-Benz", "A-Class"): "Sedan", ("Mercedes-Benz", "B-class"): "MUV/MPV",
    ("Mercedes-Benz", "C-Class"): "Sedan", ("Mercedes-Benz", "C-Coupe"): "Coupe",
    ("Mercedes-Benz", "CLA"): "Coupe", ("Mercedes-Benz", "CLS"): "Coupe",
    ("Mercedes-Benz", "E-Class"): "Sedan", ("Mercedes-Benz", "GL-Class"): "SUV",
    ("Mercedes-Benz", "GLA"): "SUV", ("Mercedes-Benz", "GLC"): "SUV", ("Mercedes-Benz", "GLE"): "SUV",
    ("Mercedes-Benz", "GLS"): "SUV", ("Mercedes-Benz", "M-Class"): "SUV",
    ("Mercedes-Benz", "R-Class"): "MUV/MPV", ("Mercedes-Benz", "S-Class"): "Sedan",
    ("Mercedes-Benz", "SLK-Class"): "Convertible", ("Mercedes-Benz", "V-Class"): "MUV/MPV",
    ("Mitsubishi", "Pajero"): "SUV",
    ("Nissan", "Magnite"): "SUV", ("Nissan", "Micra"): "Hatchback", ("Nissan", "Sunny"): "Sedan",
    ("Nissan", "Teana"): "Sedan", ("Nissan", "Terrano"): "SUV",
    ("Porsche", "718"): "Sports Car", ("Porsche", "911"): "Sports Car", ("Porsche", "Cayenne"): "SUV",
    ("Porsche", "Macan"): "SUV", ("Porsche", "Panamera"): "Sedan",
    ("Renault", "Duster"): "SUV", ("Renault", "Fluence"): "Sedan", ("Renault", "Kiger"): "SUV",
    ("Renault", "Kwid"): "Hatchback", ("Renault", "Pulse"): "Hatchback",
    ("Rolls-Royce", "Ghost"): "Sedan",
    ("Skoda", "Fabia"): "Hatchback", ("Skoda", "Kodiaq"): "SUV", ("Skoda", "Kushaq"): "SUV",
    ("Skoda", "Octavia"): "Sedan", ("Skoda", "Rapid"): "Sedan", ("Skoda", "Superb"): "Sedan",
    ("Ssangyong", "Rexton"): "SUV",
    ("Tata", "Altroz"): "Hatchback", ("Tata", "Grande"): "Hatchback", ("Tata", "Harrier"): "SUV",
    ("Tata", "Hexa"): "SUV", ("Tata", "Manza"): "Sedan", ("Tata", "Nano"): "Hatchback",
    ("Tata", "Nexon"): "SUV", ("Tata", "Punch"): "SUV", ("Tata", "Safari"): "SUV",
    ("Tata", "Tiago"): "Hatchback", ("Tata", "Tigor"): "Sedan", ("Tata", "Zest"): "Sedan",
    ("Toyota", "Camry"): "Sedan", ("Toyota", "Commuter"): "MUV/MPV", ("Toyota", "Corolla"): "Sedan",
    ("Toyota", "Fortuner"): "SUV", ("Toyota", "Glanza"): "Hatchback", ("Toyota", "Innova"): "MUV/MPV",
    ("Toyota", "Urban"): "SUV", ("Toyota", "Vellfire"): "MUV/MPV", ("Toyota", "Yaris"): "Sedan",
    ("Volkswagen", "Ameo"): "Sedan", ("Volkswagen", "Jetta"): "Sedan", ("Volkswagen", "Polo"): "Hatchback",
    ("Volkswagen", "Taigun"): "SUV", ("Volkswagen", "Tiguan"): "SUV", ("Volkswagen", "Vento"): "Sedan",
    ("Volvo", "S60"): "Sedan", ("Volvo", "S90"): "Sedan", ("Volvo", "V40"): "Hatchback",
    ("Volvo", "XC40"): "SUV", ("Volvo", "XC60"): "SUV", ("Volvo", "XC90"): "SUV",
}


def infer_body_type(make: str, model_full: str) -> str:
    """Look up a body type from the make + first word(s) of the model name."""
    make = (make or "").strip()
    model_full = (model_full or "").strip()
    base_model = model_full.split()[0] if model_full else ""

    # Toyota's "Etios" line splits into a sedan (Etios) and a hatchback (Etios Liva)
    if make == "Toyota" and base_model == "Etios":
        return "Hatchback" if "Liva" in model_full else "Sedan"

    return CAR_BODY_TYPES.get((make, base_model), "Unknown")


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_from_csv_text(text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(text))
    return [
        {key: _normalize_value(value) for key, value in row.items() if key is not None}
        for row in reader
    ]


def _rows_from_xlsx(path: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [
        _normalize_value(cell) or f"column_{index}"
        for index, cell in enumerate(rows[0])
    ]

    result = []
    for row in rows[1:]:
        values = [
            _normalize_value(cell)
            for cell in row
        ]
        normalized = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
        }
        result.append(normalized)
    return result


def load_rows_from_path(file_path: str | os.PathLike[str]) -> list[dict[str, str]]:
    path = str(file_path)

    if os.path.isdir(path):
        collected_rows: list[dict[str, str]] = []
        for child in sorted(Path(path).rglob("*")):
            if child.is_file() and child.suffix.lower() in SUPPORTED_EXTENSIONS:
                collected_rows.extend(load_rows_from_path(str(child)))
        return collected_rows

    if not os.path.exists(path):
        raise FileNotFoundError(f"Data file not found at: {path}")

    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        with open(path, mode="r", encoding="utf-8-sig", newline="") as handle:
            return _rows_from_csv_text(handle.read())

    if suffix == ".zip":
        with zipfile.ZipFile(path) as archive:
            for member_name in archive.namelist():
                lowered = member_name.lower()
                if lowered.endswith(".csv"):
                    return _rows_from_csv_text(archive.read(member_name).decode("utf-8-sig", errors="replace"))
                if lowered.endswith(".xlsx"):
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
                        handle.write(archive.read(member_name))
                        temp_path = handle.name
                    try:
                        return load_rows_from_path(temp_path)
                    finally:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
        return []

    if suffix in {".xlsx", ".xls"}:
        return _rows_from_xlsx(path)

    raise ValueError(f"Unsupported file type for ingestion: {path}")


def _build_source_id(file_path: str, metadata: dict[str, Any] | None = None) -> str:
    source_url = metadata.get("source_url") if metadata else None
    if source_url:
        return f"url::{source_url}"
    return f"path::{os.path.abspath(file_path)}"


def _source_already_ingested(collection: Any, source_id: str) -> bool:
    try:
        results = collection.get(where={"source_id": source_id}, include=["metadatas"], limit=1)
        return bool(results.get("ids"))
    except Exception:
        return False


def ingest_structured_csv(file_path: str, collection_name: str = "car_specs", metadata: dict[str, Any] | None = None):
    """
    Transforms structured dataset entries into readable paragraphs
    to preserve data integrity for embedding similarity matches.
    """
    if not file_path:
        print("[-] No data path provided")
        return

    try:
        rows = load_rows_from_path(file_path)
    except FileNotFoundError as exc:
        print(f"[-] Data file not found at: {file_path}")
        print(exc)
        return
    except ValueError as exc:
        print(f"[-] Unsupported ingestion target: {file_path}")
        print(exc)
        return

    client = get_chroma_client()
    collection = client.get_or_create_collection(name=collection_name)
    source_id = _build_source_id(file_path, metadata)

    if _source_already_ingested(collection, source_id):
        print(f"[*] Source already ingested, skipping: {file_path}")
        return

    documents = []
    metadatas = []
    ids = []

    print(f"[*] Processing data mapping from: {file_path}")
    for idx, row in enumerate(rows):
        description_sentences = []
        for key, value in row.items():
            if value:
                description_sentences.append(f"{key.replace('_', ' ').title()}: {value}")

        row_text = " | ".join(description_sentences)
        if not row_text:
            continue

        row_id = (
            row.get("id")
            or row.get("model_id")
            or row.get("vehicle_id")
            or row.get("test_vehicle_id")
            or f"row_{idx}"
        )

        documents.append(row_text)
        metadata_entry = {
            "source_file": os.path.basename(str(file_path)),
            "source_id": source_id,
            "row_index": idx,
        }
        if metadata:
            metadata_entry.update(metadata)
        metadatas.append(metadata_entry)
        ids.append(f"csv_{collection_name}_{row_id}")

    if documents:
        collection.add(documents=documents, metadatas=metadatas, ids=ids)
        print(f"[+] Successfully integrated {len(documents)} structured items into '{collection_name}' collection.")


def _format_car_row(row: dict[str, str]) -> tuple[str, str]:
    """
    Renders one used-car listing as a natural-language paragraph instead of a
    raw key:value dump, and inlines an inferred body type so semantic search
    can actually match "SUV"/"sedan"/etc. queries against it.

    Returns (document_text, body_type).
    """
    make = row.get("Make", "").strip()
    model = row.get("Model", "").strip()
    body_type = infer_body_type(make, model)
    # Acronym-style body types (SUV, MUV/MPV) keep their casing; plain words get lowercased
    body_label = body_type if body_type in {"SUV", "MUV/MPV"} else body_type.lower()
    body_phrase = f"a {body_label}" if body_type != "Unknown" else "a vehicle"

    year = row.get("Year", "")
    price = row.get("Price", "")
    km = row.get("Kilometer", "")
    fuel = row.get("Fuel Type", "")
    transmission = row.get("Transmission", "")
    location = row.get("Location", "")
    color = row.get("Color", "")
    owner = row.get("Owner", "")
    seller_type = row.get("Seller Type", "")
    engine = row.get("Engine", "")
    max_power = row.get("Max Power", "")
    max_torque = row.get("Max Torque", "")
    drivetrain = row.get("Drivetrain", "")
    seats = row.get("Seating Capacity", "")
    fuel_tank = row.get("Fuel Tank Capacity", "")

    sentences = [
        f"This is a used {year} {make} {model}, {body_phrase}, listed in {location} for Rs. {price}."
        if year and make and model
        else None,
        f"It has been driven {km} km, runs on {fuel} with a {transmission} transmission, "
        f"and is being sold by its {owner.lower()} owner through a {seller_type.lower()} seller."
        if km and fuel and transmission
        else None,
        f"It comes in {color} and is powered by a {engine} engine producing {max_power} "
        f"and {max_torque} of torque, with {drivetrain} drivetrain."
        if engine and max_power
        else None,
        f"It seats {seats} people and has a {fuel_tank}L fuel tank."
        if seats and fuel_tank
        else None,
    ]

    document_text = " ".join(s for s in sentences if s)
    return document_text, body_type


def ingest_local_car_specs(csv_path: str | None = None, collection_name: str = "car_specs"):
    """
    Ingests the local used-car dataset into ChromaDB as enriched natural-language
    paragraphs (with an inferred body type) rather than raw pipe-delimited fields.
    """
    target_path = csv_path or DEFAULT_LOCAL_CAR_SPECS_CSV
    print(f"[*] Ingesting local car specs from: {target_path}")

    try:
        rows = load_rows_from_path(target_path)
    except FileNotFoundError as exc:
        print(f"[-] Data file not found at: {target_path}")
        print(exc)
        return
    except ValueError as exc:
        print(f"[-] Unsupported ingestion target: {target_path}")
        print(exc)
        return

    client = get_chroma_client()
    collection = client.get_or_create_collection(name=collection_name)
    source_id = _build_source_id(target_path, {"source_type": "local_csv"})

    if _source_already_ingested(collection, source_id):
        print(f"[*] Source already ingested, skipping: {target_path}")
        return

    documents, metadatas, ids = [], [], []
    body_type_counts: dict[str, int] = {}

    for idx, row in enumerate(rows):
        document_text, body_type = _format_car_row(row)
        if not document_text:
            continue

        body_type_counts[body_type] = body_type_counts.get(body_type, 0) + 1

        documents.append(document_text)
        metadatas.append(
            {
                "source_file": os.path.basename(target_path),
                "source_type": "local_csv",
                "source_id": source_id,
                "row_index": idx,
                "make": row.get("Make", ""),
                "model": row.get("Model", ""),
                "body_type": body_type,
                "price": row.get("Price", ""),
                "year": row.get("Year", ""),
            }
        )
        ids.append(f"csv_{collection_name}_row_{idx}")

    if documents:
        collection.add(documents=documents, metadatas=metadatas, ids=ids)
        print(f"[+] Successfully integrated {len(documents)} enriched car listings into '{collection_name}' collection.")
        print(f"[+] Body type breakdown: {body_type_counts}")


if __name__ == "__main__":
    ingest_local_car_specs()